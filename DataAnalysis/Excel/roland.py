import openpyxl
from tkinter import messagebox
from tkinter import filedialog
import datetime
import time
import tkinter as tk





root = tk.Tk()
root.withdraw()

now = datetime.datetime.now()
today = (str(now.month) + "/" + str(now.day) + "/" + str(now.year))

messagebox.showinfo("Choose File", "Choose GLC Master")
file_path = filedialog.askopenfilename()
helloFile = open(file_path, "r")

wb = openpyxl.load_workbook(file_path)    # This spreadsheet contains all the identifiers I need
sheet = wb.get_sheet_by_name("Residential")
otherThing = sheet.max_row

messagebox.showinfo("Choose File", "Choose IOP")
file_path2 = filedialog.askopenfilename()
helloFile2 = open(file_path2, "r")

wb2 = openpyxl.load_workbook(file_path2)    # This spreadsheet contains all the identifiers I need
sheet2 = wb2.get_sheet_by_name("OUTPATIENT")
otherThing2 = sheet2.max_row

messagebox.showinfo("Choose File", "Choose Droege")
file_path3 = filedialog.askopenfilename()
helloFile3 = open(file_path3, "r")

wb3 = openpyxl.load_workbook(file_path3)    # This spreadsheet contains all the identifiers I need
sheet3 = wb3.get_sheet_by_name("Residential")
otherThing3 = sheet3.max_row


messagebox.showinfo("Choose File", "Choose Detox")
file_path4 = filedialog.askopenfilename()
helloFile4 = open(file_path4, "r")

wb4 = openpyxl.load_workbook(file_path4)    # This spreadsheet contains all the identifiers I need
sheet4 = wb4.get_sheet_by_name("DETOX")
otherThing4 = sheet4.max_row

messagebox.showinfo("Choose File", "Choose WRAP")
file_path5 = filedialog.askopenfilename()
helloFile5 = open(file_path5, "r")

wb5 = openpyxl.load_workbook(file_path5)    # This spreadsheet contains all the identifiers I need
sheet5 = wb5.get_sheet_by_name("Residential")
otherThing5 = sheet5.max_row


glc_emptyList = []
glc_asaList = []
glc_completedList = []
glc_arrestedList = []
glc_dischargedList = []


do_emptyList = []
do_asaList = []
do_completedList = []
do_arrestedList = []
do_dischargedList = []


dr_emptyList = []
dr_asaList = []
dr_completedList = []
dr_arrestedList = []
dr_dischargedList = []


de_emptyList = []
de_asaList = []
de_completedList = []
de_arrestedList = []
de_dischargedList = []


emptyList = []
asaList = []
completedList = []
arrestedList = []
dischargedList = []


def glcNumbers():
    for columnOfCellObjects in sheet["J134": "J155"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                glc_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                glc_emptyList.append(cellObj.value)
            elif cellObj.value == "Completed":
                glc_emptyList.append(cellObj.value)

    for i in glc_emptyList:
        if i == "ASA":
            glc_asaList.append(i)
    print(" ")
    print("ASA percentage for GLC:" + str(len(glc_asaList) / len(glc_emptyList) * 100) + str("%"))

    for i in glc_emptyList:
        if i == "Discharged":
            glc_dischargedList.append(i)
    print("Discharged percentage for GLC:" + str(len(glc_dischargedList) / len(glc_emptyList) * 100) + str("%"))

    for i in glc_emptyList:
        if i == "Completed":
            glc_completedList.append(i)
    print("Graduated percentage for GLC:" + str(len(glc_completedList) / len(glc_emptyList) * 100) + str("%"))


def doughertyNumbers():
    for columnOfCellObjects in sheet2["I93": "I98"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "GRAD":
                do_emptyList.append(cellObj.value)
            elif cellObj.value == "Unsucc":
                do_emptyList.append(cellObj.value)

    for i in do_emptyList:
        if i == "GRAD":
            do_completedList.append(i)
    print(" ")
    print("Graduated percentage for IOP:" + str(len(do_completedList) / len(do_emptyList) * 100) + str("%"))

    for i in do_emptyList:
        if i == "Unsucc":
            do_dischargedList.append(i)
    print("Unsuccessful percentage for IOP:" + str(len(do_dischargedList) / len(do_emptyList) * 100) + str("%"))


def droegeNumbers():
    for columnOfCellObjects in sheet3["I58": "I83"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Graduate":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Complete":
                dr_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                dr_emptyList.append(cellObj.value)

    for i in dr_emptyList:
        if i == "ASA":
            dr_asaList.append(i)
    print(" ")
    print("ASA percentage for Droege:" + str(len(dr_asaList) / len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Graduate":
            dr_completedList.append(i)
    print("Graduated percentage for Droege:" + str(len(dr_completedList) / len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Complete":
            dr_arrestedList.append(i)
    print("Completion percentage for Droege:" + str(len(dr_arrestedList) / len(dr_emptyList) * 100) + str("%"))

    for i in dr_emptyList:
        if i == "Discharged":
            dr_dischargedList.append(i)
    print("Discharged percentage for Droege:" + str(len(dr_dischargedList) / len(dr_emptyList) * 100) + str("%"))


def detoxNumbers():
    for columnOfCellObjects in sheet4["I25": "I113"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "Med Discharge":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "comp":
                de_emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                de_emptyList.append(cellObj.value)

    for i in de_emptyList:
        if i == "ASA":
            de_asaList.append(i)
    print(" ")
    print("ASA percentage for Detox:" + str(len(de_asaList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "Med Discharge":
            de_completedList.append(i)
    print("MEd Discharge percentage for Detox:" + str(len(de_completedList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "comp":
            de_arrestedList.append(i)
    print("Completion percentage for Detox:" + str(len(de_arrestedList) / len(de_emptyList) * 100) + str("%"))

    for i in de_emptyList:
        if i == "Discharged":
            de_dischargedList.append(i)
    print("Discharged percentage for Detox:" + str(len(de_dischargedList) / len(de_emptyList) * 100) + str("%"))


def wrapNumbers():
    for columnOfCellObjects in sheet5["H55": "H84"]:
        for cellObj in columnOfCellObjects:
            if cellObj.value == "ASA":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Completed":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Arrested":
                emptyList.append(cellObj.value)
            elif cellObj.value == "Discharged":
                emptyList.append(cellObj.value)

    for i in emptyList:
        if i == "ASA":
            asaList.append(i)
    print(" ")
    print("ASA percentage for WRAP:" + str(len(asaList) / len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Completed":
            completedList.append(i)
    print("Completed percentage for WRAP:" + str(len(completedList) / len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Arrested":
            arrestedList.append(i)
    print("Arrested percentage for WRAP:" + str(len(arrestedList) / len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Discharged":
            dischargedList.append(i)
    print("Discharged percentage for WRAP:" + str(len(dischargedList) / len(emptyList) * 100) + str("%"))

    for i in emptyList:
        if i == "Un Succ Dis":
            dischargedList.append(i)
    print(
        "Un Successful Discharge percentage for WRAP:" + str(len(dischargedList) / len(emptyList) * 100) + str("%"))


def gO_Now():
    glcNumbers()
    droegeNumbers()
    detoxNumbers()
    wrapNumbers()
    doughertyNumbers()
    print(today)
    time.sleep(3)
    exit()

gO_Now()
